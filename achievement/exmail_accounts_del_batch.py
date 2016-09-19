#-*- coding=utf-8 -*-
__author__ = 'ajioy'
__create_data__ = '2016/8/29 17:22'
__py_version__ = '2.7.8'
__description__ = '腾讯企业邮箱批量离职程序，源表有两个文件，一个数据仓库，'\
                    '一个离职人员清单原表，输出一个结果表显示执行状态 '\
                    '知识点：读写excel、txt，OAuth2.0、程序计时'

#   注：openpyxl 需要pip安装
from openpyxl import load_workbook #for read,modify excel sheet
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import RED,GREEN,BLUE,DARKYELLOW
import time
import datetime
import TecentExmailOperator


# 原表字段所在列,手动配置，可读入conf.txt
filed_origin = {
    'alias':7,
    'id':2,
    'job':6,
    'status':9,
    'name':3,
    'date':8
}

# 仓库字段所在列，腾讯已定义，固定格式，但也许会变
filed_house = {
    'extid':8,
    'email':2,
    'name':1
}

# 统计各种结果状态的个数
count = {
    'cnt_suc':0,
    'cnt_fail':0,
    'cnt_nfound':0,
    'cnt_ndchk':0,
    'cnt_date':0
}

# 封装类，对excel的基本操作进行封装，最后注意，保存方法需要手动调用
class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
        self.rows = len(self.ws.rows)
        self.cols = len(self.ws.columns)

    def write(self, r, c, val):
        self.ws.cell(row=r, column=c).value = val
        #self.wb.save(self.filename)

    def merge(self, rangstring):
        self.ws.merge_cells(rangstring)
        self.wb.save(self.filename)

    def cellstyle(self, r, c, font,\
        align = Alignment(horizontal='left', vertical='center')):
        cell = self.ws.cell(row=r ,column=c)
        cell.font = font
        cell.alignment = align

    def getnRows(self): # actived sheet rows
        return self.rows

    def getnCols(self):  #actived sheet cols
        return self.cols

    def getCellValue(self,nrow, ncol, bisdate=False): # 取值时去空格或Tab制表符
        val = self.ws.cell(row = nrow, column = ncol).value
        if val != None:
            if bisdate:
                return val
            else:
                val = val.lstrip() # 去左边
                val = val.rstrip() # 去右边
                return val
        else:
            return None

    def setmark(self, r, c, val, colors):
        self.ws.cell(row=r, column=c).value = val
        font = Font(name=u'宋体', size=12, color=colors, bold=False)
        self.cellstyle(r, c,font) # self.cellstyle(self,r,c,font) wrong
        #self.wb.save(self.filename)

    def save(self): #   手动保存，write()\setmark()的save方法不一一执行，而是最后统一保存，效率提升16倍
        self.wb.save(self.filename)

    def saveas(self, filename):
        self.wb.save(filename)

# 获取字段所在行
def getCurRow(wb):

    for row in range(1,5):
        for col in range(1,5):
            if wb.getCellValue(row, col) == u'工号':
                return row


#根据邮箱名删除账号，成功标记successful,失败标记failed，累加个数
def delAccount(Alias, wr_origin, nCurRow):

    op = TecentExmailOperator.TecentExmailOperator()
    # 慎用，一旦删除将无法在系统日志模块中恢复，除非你账号无误
    # 建议先禁用，然后管理员手动删除，可在7天内恢复
    res = op.delAccountByAlias(Alias)

    if res == 'done':
        strCur = ('successful', BLUE)
        count['cnt_suc'] += 1
    elif res == 'user_not_found':
        strCur = ('failed delete', RED)
        count['cnt_fail'] += 1
    else:
         strCur = ('failed delete', RED)
         count['cnt_fail'] += 1

    wr_origin.setmark(nCurRow, filed_origin['status'], strCur[0], strCur[1])
    wr_origin.setmark(nCurRow, filed_origin['status'] + 2, Alias, BLUE)


def printResult():
    print "---------------------------------------------------------------"
    print "total handle:%d\r" % (count['cnt_suc'] + count['cnt_fail'] + \
                                 count['cnt_ndchk'] + count['cnt_nfound'] + count['cnt_date'])
    print "id need check:%d\r" % count['cnt_ndchk']
    print "attention date:%d\r" % count['cnt_date']
    print "successful:%d\r" % count['cnt_suc']
    print "failed:%d\r" % count['cnt_fail']
    print "not found:%d\r" % count['cnt_nfound']
    print "---------------------------------------------------------------"


def main():
    print "start handling..."

    origin_file = u'离职人员标准表.xlsx'
    wr_origin = Write_excel(origin_file)

    house_file = 'address_biz.xlsx'
    wr_house = Write_excel(house_file)

    house_nRows = wr_house.getnRows() + 1   # 注意边界,不加1会遗漏最后一个
    origin_nRows = wr_origin.getnRows()
    origin_ncurRow = getCurRow(wr_origin)  # 当前行游标，初始指向原表字段行
    nError = 0
    # 逐行读入原表数据行
    while origin_ncurRow < origin_nRows:
        origin_ncurRow += 1 # 字段行+1为数据处理行
        st_job = wr_origin.getCellValue(origin_ncurRow, filed_origin['job'])

        if st_job == u'客户经理':   # 筛选掉客户经理，不处理,标记no_account-->改成not_found，视作找不到，用户可无视
            wr_origin.setmark(origin_ncurRow, filed_origin['status'],'not_found', DARKYELLOW)
            count['cnt_nfound']+=1
            continue

        origin_date = wr_origin.getCellValue(origin_ncurRow, filed_origin['date'], True)
        origin_name = wr_origin.getCellValue(origin_ncurRow, filed_origin['name'])
        origin_id = wr_origin.getCellValue(origin_ncurRow, filed_origin['id'])

        if origin_date == None:
            if origin_id == None and origin_name == None:  # 姓名和工号都是空是最后一行，跳出
                break
            else:
                wr_origin.setmark(origin_ncurRow, filed_origin['status'], 'not_found', DARKYELLOW)
                count['cnt_nfound'] += 1
                continue

        #   print origin_ncurRow
        origin_date = origin_date.strftime('%m-%d-%y')  # 注意，日期如果是文本或其他格式会报错
        curDate = datetime.datetime.now().strftime('%m-%d-%y')
        if origin_date > curDate:  # 离职日期还未到   日期可直接比较，datetime重载运算符
            wr_origin.setmark(origin_ncurRow, filed_origin['status'], 'attention date', RED)
            count['cnt_date'] += 1
            continue

        bIf_found = False   # 仓库姓名是否找到
        bIf_idequal = False # 仓库编号是否相等
        for cur in range(2, house_nRows): # 查询原表姓名是否出现在仓库中
            house_name = wr_house.getCellValue(cur, filed_house['name'])
            if origin_name == house_name:   # 仓库中存在此姓名
                bIf_found = True
                extId = wr_house.getCellValue(cur, filed_house['extid'])
                alias = wr_house.getCellValue(cur, filed_house['email'])
                if origin_id == extId: # 原表工号与仓库编号相同
                    delAccount(alias, wr_origin, origin_ncurRow)
                    bIf_idequal = True
                    break   # 跳出for循环
                else:   # 原表工号与仓库编号不相等的情况下，判断原表邮箱是否与仓库邮箱相等
                    origin_alias = wr_origin.getCellValue(origin_ncurRow, filed_origin['alias'])
                    if alias == origin_alias:
                        delAccount(alias, wr_origin, origin_ncurRow)
                        bIf_idequal = True
                        break  # 跳出for循环
                    continue

        #   遍历完仓库后未发现原表姓名，标记not_found
        if bIf_found == True:
            if bIf_idequal == False:    # 仓库中没有对应的编号
                wr_origin.setmark(origin_ncurRow, filed_origin['status'], 'id_need_check', GREEN)
                count['cnt_ndchk']+=1
        else:
            wr_origin.setmark(origin_ncurRow, filed_origin['status'], 'not_found', DARKYELLOW)
            count['cnt_nfound'] += 1


    #   表格处理完后一次性保存
    wr_origin.saveas(u"output.xlsx")

    printResult()




if __name__ == '__main__':
    starttime = time.clock()
    main()
    endtime = time.clock()
    print "successful excuted %f" % (endtime-starttime)
