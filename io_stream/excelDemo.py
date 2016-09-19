#-*- coding=utf-8 -*-
__author__ = 'ajioy'
__create_data__ = '2016/8/26 12:56'
__py_version__ = '2.7.8'
__description__ = ''

import csv
from openpyxl import load_workbook #for read,modify
from openpyxl import Workbook  #for write
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import RED,GREEN,BLUE,YELLOW
from time import ctime

# 原表字段
filed_origin = {
    'alias':7,
    'id':2,
    'job':6,
    'status':9,
    'name':3
}

# 仓库字段
filed_house = {
    'extid':8,
    'email':2,
    'name':1
}


class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
        self.rows = len(self.ws.rows)
        self.cols = len(self.ws.columns)

    def write(self, r, c, val):
        self.ws.cell(row=r, column=c).value = val
        self.wb.save(self.filename)

    def merge(self, rangstring):
        self.ws.merge_cells(rangstring)
        self.wb.save(self.filename)

    def cellstyle(self, r, c, font,\
        align = Alignment(horizontal='left', vertical='center')):
        cell = self.ws.cell(row=r ,column=c)
        cell.font = font
        cell.alignment = align

    def getnRows(self): # actived sheet rows
        return self.rows

    def getnCols(self):  #actived sheet cols
        return self.cols

    def getCellValue(self,nrow, ncol):
        return self.ws.cell(row = nrow, column = ncol).value

    def setmark(self, r, c, val, colors):
        self.ws.cell(row=r, column=c).value = val
        font = Font(name=u'宋体', size=12, color=colors, bold=True)
        self.cellstyle(r, c,font) # self.cellstyle(self,r,c,font) wrong
        self.wb.save(self.filename)



def trimlr(val):
    if val != None:
        val = val.lstrip()  # 去左边空格
        val = val.rstrip()  # 去右边空格
        return val
    else:
        return None

def delAccount(alias):
    return True

def handelNoAliasFind():    # 没有邮箱字段处理
    pass


def main():
    # 逐行读入原表，判断是否有邮箱
    origin = u'标准离职表0830.xlsx'
    print origin
    wr_origin = Write_excel(origin)
    house = 'house.xlsx'
    wr_house = Write_excel(house)
    nHouse = wr_house.getnRows() + 1

    for rows in range(2,20):    # 第1行为保留字段行
        st_job = wr_origin.getCellValue(rows, filed_origin['job'])
        st_job = trimlr(st_job)

        if st_job == u'客户经理':   # 筛选掉客户经理，不处理
            wr_origin.setmark(rows, filed_origin['status'],'saler_no_account', GREEN)
            continue

        st_alias = wr_origin.getCellValue(rows,filed_origin['alias'])
        st_alias = trimlr(st_alias)
        if st_alias == None:   # 没有邮箱字段
            print "no alias"

            sId = wr_origin.getCellValue(rows,filed_origin['id'])
            sId = trimlr(sId)
            for each in range(1,nHouse):    # 根据原表工号，匹配仓库编号
               extId =  wr_house.getCellValue(each, filed_house['extid'])
               if extId != None:
                   extId = trimlr(extId)
                   if sId == extId: # 匹配到了，则取邮箱号
                       email = wr_house.getCellValue(each, filed_house['email'])
                       if delAccount(email) == True:
                            wr_origin.setmark(rows, filed_origin['status'],'successful', GREEN)
                       else:
                           wr_origin.setmark(rows, filed_origin['status'], 'delete failed', RED)
             # 没有匹配到工号，则查询原表姓名是否出现，出现则加入人工核查，没有则标记no_reacord

        else:   # 有邮箱字段
            st_name = wr_origin.getCellValue(rows, filed_origin['name'])
            st_name = trimlr(st_name)
            # 容错处理
            # 原表姓名是否存在于仓库
            bOccurflag = False
            for each in range(2, nHouse):
                st_house_name = wr_house.getCellValue(each, filed_house['name'])
                #if type(st_house_name) != type(str):    # 此句无法判断表格内容类型是否为字符串
                    #print "not right type, string"
                 #   continue
                try:
                    st_house_name = trimlr(st_house_name)
                    if  st_name == st_house_name: # 姓名存在
                        print "存在"
                        email = wr_house.getCellValue(each, filed_house['email'])
                        if delAccount(email) == True:
                            wr_origin.setmark(rows, filed_origin['status'], 'successful', GREEN)
                        else:
                            wr_origin.setmark(rows, filed_origin['status'], 'delete failed', RED)
                        bOccurflag = True
                        break
                    else:   # 不存在，则标记no_record_in_addressbiz
                        pass

                except AttributeError,e:
                    print e,"请查阅仓库address中姓名字段是否有数字开头"


            if not bOccurflag:  # 仓库中没有对应的姓名
                wr_origin.setmark(rows, filed_origin['status'], 'no_record', RED)





if __name__ == '__main__':
    main()
    print "ok.all done!"



'''
    print "starting at:",ctime()
    wr = Write_excel('')
    wr.write('J2','ajioy modified')
    #   设置单元格样式
    # font = Font(name = u'微软雅黑', size = 20, color = RED, bold = True)
    # align = Alignment(horizontal='center', vertical='center')
    # wr.cellstyle('A12', font, align)
    # wr.write('A12', 'ajioy modified')
    print wr.getRows(), wr.getCols()
    print "all done at:",ctime()
'''


# #read begin
# wb = load_workbook(r'demo.xlsx')
# sheets = wb.get_sheet_names()
# sheet0 = sheets[0]
# ws = wb.get_sheet_by_name('demo')
#
# rows = ws.rows
# cols = ws.columns
#
# content = []
# for row in rows:
#     line = [col.value for col in row]
#     content.append(line)
#
# print ws.cell('B2').value
# print ws.cell(row=12, column=2).value
# #read end
#
# #write begin
# wb = Workbook()
# ws = wb.get_active_sheet()
# print ws.title
# ws.title = 'New Title'
#
# ws.cell('A1').value = 'ajioy'
# ws.cell(row = 2, column = 1).value = 26
#
# new_ws = wb.create_sheet(title = 'new_sheet')
# for row in range(1,100):  #row < 1 or col < 1 will receive a warning
#     for col in range(1,10):
#         new_ws.cell(row=row, column=col).value = row+col
#
# wb.save(r'helloajioy.xlsx')
# #write end
#
# i = 0
# with open('address_biz.csv','rb') as f:
#     reader = csv.reader(f)  # tuple method
#     headers = next(reader)
#     for row in reader:
#         #s = str(row).decode('gb2312') #wrong
#         #s = ', '.join(row).decode('gbk')  # 部分正常
#         s = ', '.join(row).decode('gbk') #正常显示中文
#         #print s
#         ss = row[7].rstrip()    #去\t制表符
#         print ss, len(ss), row[1]
#         i+=1
#     print i
#         #print row
    # reader = csv.reader(f, delimiter='\t')
    # headings = next(reader)
    # for r in reader:
    #     s = ', '.join(r).decode('gb2312')
    #     ss =
    #     print ss, len(ss)

# with open('exportcsv.csv', 'wb') as f:
#     csvwt = csv.writer(f)
#     for row in range(15):
#         s1 = u'人才'.encode('gb2312') #write gbk
#         s2 = len('A4881')
#         csvwt.writerow([s1] * 5 + ['我宝贝儿'])
#         csvwt.writerow(['study','python','is','A4881',s2])



# import xlrd
# import xlwt
#
# workbook = xlrd.open_workbook(r"demo.xlsx")
# print workbook.sheet_names()
#
# curSheet = workbook.sheet_by_index(0)
# print curSheet.name, curSheet.nrows,curSheet.ncols
#
# rows = curSheet.row_values(3)
#
# cell = curSheet.cell_value(1,7)
# size = len(cell)
# cell2 = cell
# #cell2 = cell2.expandtabs()
# cell2 = cell2.rstrip() #去/t制表符、空格之类
# size2 = len(cell2)
# cell3 = curSheet.cell_value(1,8)
#
# if cell == cell3:
#     print "cell == cell3"
# elif cell2 == cell3:
#     print "cell2 == cell3"
#
# cell4 = curSheet.cell_value(2,7)
# size3 = len(cell4)
#
# #cell = curSheet.cell(2,1).value.encode('utf-8')
# print cell,size,size2,size3
