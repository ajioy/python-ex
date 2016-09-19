# 封装类，对excel的基本操作进行封装，最后注意，保存方法需要手动调用
from openpyxl import load_workbook #for read,modify excel sheet
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import RED,GREEN,BLUE,DARKYELLOW


class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
        self.rows = len(self.ws.rows)
        self.cols = len(self.ws.columns)

    def write(self, r, c, val):
        self.ws.cell(row=r, column=c).value = val
        #self.wb.save(self.filename)

    def merge(self, rangstring):
        self.ws.merge_cells(rangstring)
        self.wb.save(self.filename)

    def cellstyle(self, r, c, font,\
        align = Alignment(horizontal='left', vertical='center')):
        cell = self.ws.cell(row=r ,column=c)
        cell.font = font
        cell.alignment = align

    def getnRows(self): # actived sheet rows
        return self.rows

    def getnCols(self):  #actived sheet cols
        return self.cols

    def getCellValue(self,nrow, ncol, bisdate=False): # 取值时去空格或Tab制表符
        val = self.ws.cell(row = nrow, column = ncol).value
        if val != None:
            if bisdate:
                return val
            else:
                val = val.lstrip() # 去左边
                val = val.rstrip() # 去右边
                return val
        else:
            return None

    def setmark(self, r, c, val, colors):
        self.ws.cell(row=r, column=c).value = val
        font = Font(name=u'宋体', size=12, color=colors, bold=False)
        self.cellstyle(r, c,font) # self.cellstyle(self,r,c,font) wrong
        #self.wb.save(self.filename)

    def save(self): #   手动保存，write()\setmark()的save方法不一一执行，而是最后统一保存，效率提升16倍
        self.wb.save(self.filename)

    def saveas(self, filename):
        self.wb.save(filename)